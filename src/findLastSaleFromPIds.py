import openpyxl

excel_file = openpyxl.load_workbook("findLastSaleFromPIds.xlsx")
spreadsheet1 = excel_file["PIds"]

# for rowIndex in range(2, spreadsheet1.max_row+1):
#     cell = spreadsheet1.cell(rowIndex, 1).value
#     print(cell)

# for rowArrContent in spreadsheet1:
for rowIndex in range(2, spreadsheet1.max_row+1):
    rowArrContent = spreadsheet1[rowIndex]
    pIds = rowArrContent[0].value
    l1Ib = rowArrContent[1].value
    pIdArray = pIds.split(",")

    for i in range(len(pIdArray)):
        if(str(pIdArray[i]) == str(l1Ib)):
            print("lastSale:", end='   ')
            print(pIdArray[i-1], end='   ')
            lastSaleCell = spreadsheet1.cell(row=rowIndex, column=4)
            lastSaleCell.value = pIdArray[i-1];
            print("lastSaleLeader:", end='   ')
            print(pIdArray[i-2], end='   ')
            lastSaleLeaderCell = spreadsheet1.cell(row=rowIndex, column=5)
            lastSaleLeaderCell.value = pIdArray[i-2];
            print('')

excel_file.save("solution.xlsx")