import openpyxl
from time import time

# The result will be same even the script runs several times.
print("Script begin:")
start = time()
# !! RESET the number of maxRow before running the script, before spreadsheet.max_row is not accurate.
excelFileName = "splitGroupAndCloneRow.xlsx"
toDoSpreadsheetName = "rebateAccList"
maxRow = 15588
cloneRowIndex = maxRow + 5
maxColumn = 24   #column index start from 1
columnIndexSplit = 3
ibGroupIndex = 10

excel_file = openpyxl.load_workbook(excelFileName)
spreadsheetTest = excel_file[toDoSpreadsheetName]
for rowIndex in range(2, maxRow+1):     #skip head row
    print(f"start row: {rowIndex}")
    rowArrContent = spreadsheetTest[rowIndex]
    ibGroups = rowArrContent[columnIndexSplit].value
    if not ibGroups == None: 
        ibGroupArray = ibGroups.split(",")
    else:
        ibGroupArray = []   # skip the row if ibGroups is null
    for ibGroupArrayIndex in range(len(ibGroupArray)):
        for columnIndex in range(1, maxColumn):
            oriCell = spreadsheetTest.cell(row=rowIndex, column=columnIndex)
            cloneCell = spreadsheetTest.cell(row=cloneRowIndex, column=columnIndex)
            cloneCell.value = oriCell.value
        ibGroupCell = spreadsheetTest.cell(row=cloneRowIndex, column=ibGroupIndex)
        ibGroupCell.value = ibGroupArray[ibGroupArrayIndex]
        cloneRowIndex = cloneRowIndex + 1

excel_file.save(excelFileName)
end = time()
print(f"Runtime of the program is {end - start}")
print("Script end!")