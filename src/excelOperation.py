import openpyxl

excel_file = openpyxl.load_workbook("test.xlsx")
spreadsheet1 = excel_file["Sheet1"]

# for rowIndex in range(2, spreadsheet1.max_row+1):
#     cell = spreadsheet1.cell(rowIndex, 1).value
#     print(cell)

for rowArrContent in spreadsheet1:
    for cell in rowArrContent:
        print(cell.value, end='   ')
    print('')